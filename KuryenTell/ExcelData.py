import openpyxl
from openpyxl import Workbook
import pathlib


def create_excel_sheet():
    """
        Create an Excel sheet with predefined column headers if it doesn't exist.

        If the 'Saved_Data.xlsx' file already exists, do nothing. Otherwise, create a new
        workbook, add a sheet, and set up the column headers.

        """
    main_xlsx_file = pathlib.Path("Saved_Data.xlsx")
    if main_xlsx_file.exists():
        pass
    else:
        main_xlsx_file = Workbook()
        a_sheet = main_xlsx_file.active
        a_sheet['A1'] = "Appliance No."
        a_sheet['B1'] = "Date Registered"
        a_sheet['C1'] = "Name"
        a_sheet['D1'] = "Brand"
        a_sheet['E1'] = "Model"
        a_sheet['F1'] = "Wattage"
        a_sheet['G1'] = "Hours Used"
        a_sheet['H1'] = "Total Bill"
        a_sheet['I1'] = "kW Hourly"
        a_sheet['J1'] = "kW Daily"
        a_sheet['K1'] = "kW Weekly"
        a_sheet['L1'] = "kW Monthly"
        a_sheet['M1'] = "kW Yearly"
        a_sheet['N1'] = "Bill Percentage"
        main_xlsx_file.save("Saved_Data.xlsx")


def registered_no(reg_num):
    """
        Update the given Tkinter IntVar with the next available Appliance No.

        Args:
            reg_num (tk.IntVar): Tkinter IntVar to be updated with the next available
            Appliance No.

        """
    main_xlsxfile = openpyxl.load_workbook('Saved_Data.xlsx')
    sheet = main_xlsxfile.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        reg_num.set(max_row_value + 1)
    except:
        reg_num.set("1")


create_excel_sheet()
